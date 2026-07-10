"""
app.py — Serveur unique du projet Hôpital 4.0
================================================
Regroupe en un seul fichier :
  - le serveur web (sert la page HTML + API REST pour le frontend "Contrôle Colis")
  - la liaison série avec l'Arduino (anciennement liaison_arduino.py)
  - la base de données SQLite (colis + historique)

Installation :
    pip install flask flask-cors pyserial openpyxl

Lancement :
    python app.py
Puis ouvrir dans le navigateur :
    http://localhost:5000
"""

import os
import sqlite3
import threading
import time
from datetime import datetime
from io import BytesIO

from flask import Flask, render_template, request, jsonify
from flask_cors import CORS

import serial
import serial.tools.list_ports
from openpyxl import load_workbook

# ============================================================
# CONFIGURATION
# ============================================================
BASE_DIR     = os.path.dirname(os.path.abspath(__file__))
DB_PATH      = os.path.join(BASE_DIR, "colis.db")
BAUDRATE     = 9600
PORT_SECOURS = "COM7"          # utilisé seulement si aucun Arduino n'est détecté automatiquement

app = Flask(__name__)
CORS(app)   # autorise les appels du frontend, y compris en origine 'null' (fichier ouvert en file://)

db_lock     = threading.Lock()   # évite les écritures SQLite concurrentes (thread série + requêtes HTTP)
serial_lock = threading.Lock()   # évite les écritures concurrentes sur le port série
arduino     = None               # objet serial.Serial une fois la connexion établie

messages         = []   # petite file en mémoire pour le polling /messages (notifications live côté site)
next_message_id  = 1


# ============================================================
# BASE DE DONNÉES
# ============================================================
def get_db():
    conn = sqlite3.connect(DB_PATH, check_same_thread=False)
    conn.row_factory = sqlite3.Row
    return conn


def init_db():
    with db_lock, get_db() as conn:
        conn.execute("""
            CREATE TABLE IF NOT EXISTS colis (
                id           INTEGER PRIMARY KEY AUTOINCREMENT,
                numero_colis TEXT    NOT NULL,
                destination  TEXT    NOT NULL,
                angle_servo  INTEGER NOT NULL,
                statut       TEXT    NOT NULL DEFAULT 'en_attente'
            )
        """)
        conn.execute("""
            CREATE TABLE IF NOT EXISTS historique (
                id      INTEGER PRIMARY KEY AUTOINCREMENT,
                message TEXT NOT NULL,
                date    TEXT NOT NULL
            )
        """)
        conn.commit()


def ajouter_historique(message):
    """Ajoute une ligne d'historique en BDD + dans la file de messages live (pour le polling)."""
    global next_message_id

    date_str = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
    with db_lock, get_db() as conn:
        conn.execute("INSERT INTO historique (message, date) VALUES (?, ?)", (message, date_str))
        conn.commit()

    messages.append({"id": next_message_id, "text": message})
    next_message_id += 1
    if len(messages) > 50:
        messages.pop(0)


# ============================================================
# LOGIQUE MÉTIER
# (fonctions utilisées à la fois par les routes HTTP et directement
#  par le thread série, sans repasser par une requête HTTP interne)
# ============================================================
def prochain_colis():
    """Cherche le plus ancien colis 'en_attente', le marque 'traite' et renvoie son angle servo."""
    with db_lock, get_db() as conn:
        row = conn.execute(
            "SELECT * FROM colis WHERE statut = 'en_attente' ORDER BY id ASC LIMIT 1"
        ).fetchone()

        if row is None:
            return {"status": "aucun_colis"}

        conn.execute("UPDATE colis SET statut = 'traite' WHERE id = ?", (row["id"],))
        conn.commit()

    ajouter_historique(
        f"📦 Colis {row['numero_colis']} trié → {row['destination']} (servo {row['angle_servo']}°)"
    )
    return {"status": "ok", "angle_servo": row["angle_servo"], "numero_colis": row["numero_colis"]}


# ============================================================
# LIAISON SÉRIE AVEC L'ARDUINO (ex liaison_arduino.py, fusionné ici)
# ============================================================
def detecter_port_arduino():
    for p in serial.tools.list_ports.comports():
        if any(mot in (p.description or "") for mot in ["Arduino", "CH340", "USB", "Genuino"]):
            return p.device
    return PORT_SECOURS


def boucle_serie():
    """Tourne dans un thread séparé : lit l'Arduino en continu et déclenche la logique métier."""
    global arduino

    port = detecter_port_arduino()
    print(f"🔌 Tentative de connexion série sur le port : {port}...")
    try:
        arduino = serial.Serial(port, BAUDRATE, timeout=1)
        time.sleep(2)  # laisse l'Arduino redémarrer après l'ouverture du port
        print(f"✅ Connexion série établie sur {port} !")
    except Exception as e:
        print(f"❌ Impossible d'ouvrir le port {port} : {e}")
        print("👉 Le site fonctionnera quand même, mais sans liaison Arduino réelle.")
        return

    print("🚀 Liaison série active ! En écoute des capteurs...")
    while True:
        try:
            ligne = arduino.readline().decode(errors="ignore").strip()
            if not ligne:
                continue

            print(f"📥 Arduino dit : '{ligne}'")
            commande = ligne.upper().replace(" ", "").replace("_", "")

            if "CAPTEUR1" in commande:
                ajouter_historique("📡 Capteur 1 franchi")
                resultat = prochain_colis()
                if resultat["status"] == "ok":
                    angle = resultat["angle_servo"]
                    print(f"🔄 Angle {angle}° -> envoi au servomoteur")
                    with serial_lock:
                        arduino.write(f"SERVO:{angle}\n".encode())
                else:
                    print("ℹ️ Aucun colis en attente.")

            elif "CAPTEUR2" in commande:
                ajouter_historique("📡 Capteur 2 franchi")

        except Exception as e:
            print(f"💥 Erreur dans la boucle série : {e}")
            time.sleep(1)


# ============================================================
# ROUTE — PAGE WEB PRINCIPALE
# ============================================================
@app.route("/")
def index():
    # Le fichier doit se trouver dans templates/hopital-4-0-final.html
    return render_template("hopital-4-0-final.html")


# ============================================================
# ROUTES — API COLIS
# ============================================================
@app.route("/api/colis")
def api_colis():
    with db_lock, get_db() as conn:
        rows = conn.execute("SELECT * FROM colis ORDER BY id ASC").fetchall()
    return jsonify({"colis": [dict(r) for r in rows]})


@app.route("/prochain-colis")
def route_prochain_colis():
    return jsonify(prochain_colis())


@app.route("/upload", methods=["POST"])
def upload():
    fichier = request.files.get("fichier")
    if not fichier:
        return jsonify({"status": "erreur", "message": "Aucun fichier reçu"}), 400

    try:
        classeur = load_workbook(filename=BytesIO(fichier.read()))
        feuille = classeur.active

        lignes_importees = 0
        with db_lock, get_db() as conn:
            # Ligne d'en-tête attendue : numero_colis | destination | angle_servo
            for ligne in feuille.iter_rows(min_row=2, values_only=True):
                if not ligne or ligne[0] is None:
                    continue
                numero_colis, destination, angle_servo = ligne[0], ligne[1], ligne[2]
                conn.execute(
                    "INSERT INTO colis (numero_colis, destination, angle_servo, statut) "
                    "VALUES (?, ?, ?, 'en_attente')",
                    (str(numero_colis), str(destination), int(angle_servo)),
                )
                lignes_importees += 1
            conn.commit()

        ajouter_historique(f"📥 Import Excel : {lignes_importees} colis ajoutés")
        return jsonify({"status": "ok", "lignes_importees": lignes_importees})

    except Exception as e:
        return jsonify({"status": "erreur", "message": str(e)}), 500


# ============================================================
# ROUTES — HISTORIQUE & MESSAGES LIVE
# ============================================================
@app.route("/api/historique")
def api_historique():
    with db_lock, get_db() as conn:
        rows = conn.execute("SELECT * FROM historique ORDER BY id DESC").fetchall()
    return jsonify({"historique": [dict(r) for r in rows]})


@app.route("/api/vider-historique", methods=["POST"])
def vider_historique():
    with db_lock, get_db() as conn:
        conn.execute("DELETE FROM historique")
        conn.commit()
    return jsonify({"status": "ok", "message": "🗑️ Historique vidé !"})


@app.route("/messages")
def route_messages():
    return jsonify({"messages": messages})


# ============================================================
# ROUTE — CAPTEUR (signal manuel depuis le site, ou capteur libre)
# ============================================================
@app.route("/capteur", methods=["POST"])
def capteur():
    data = request.get_json(silent=True) or {}
    message = (data.get("message") or "").strip()
    if not message:
        return jsonify({"status": "erreur", "message": "Message vide"}), 400

    ajouter_historique(f"📡 {message}")
    return jsonify({"status": "ok"})


# ============================================================
# DÉMARRAGE
# ============================================================
if __name__ == "__main__":
    init_db()

    # Thread dédié à la liaison série : ne bloque jamais le serveur web
    thread_serie = threading.Thread(target=boucle_serie, daemon=True)
    thread_serie.start()

    print("🌐 Serveur Flask démarré sur http://localhost:5000")
    # use_reloader=False : indispensable, sinon Flask redémarre le process en mode debug
    # et tente d'ouvrir le port série DEUX fois en même temps -> erreur.
    app.run(host="0.0.0.0", port=5000, debug=True, use_reloader=False)
